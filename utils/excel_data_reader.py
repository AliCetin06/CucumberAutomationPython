import openpyxl
from openpyxl import Workbook
from openpyxl.worksheet.worksheet import Worksheet


class ExcelDataReader:
    _workbook: Workbook = None
    _sheet: Worksheet = None

    @classmethod
    def init_excel_data(cls, file_path: str = "/Users/alicetin/Desktop/eclipse/CucumberAutomation/src/test/resources/data/TestData.xlsx") -> None:
        """Excel dosyasini yükler ve ilk çalişma sayfasini (sheet) seçer."""
        cls._workbook = openpyxl.load_workbook(file_path, data_only=True)
        # Ilk sayfayi seçer (Java'daki getSheetAt(0) karşiliği)
        cls._sheet = cls._workbook.worksheets[0]

    @classmethod
    def get_data(cls, row_number: int, column_number: int) -> str:
        """
        Belirtilen satir ve sütundaki hücre değerini string olarak döndürür.
        Not: openpyxl indeksleri 1'den başlar (1-based index).
        Java uyumluluğu için 0-based index gelirse +1 ekliyoruz.
        """
        cell_value = cls._sheet.cell(row=row_number + 1, column=column_number + 1).value
        return str(cell_value) if cell_value is not None else ""

    @classmethod
    def get_total_row(cls) -> int:
        """Toplam dolu satir sayisini döndürür."""
        return cls._sheet.max_row

    @classmethod
    def get_total_column(cls) -> int:
        """Ilk satirdaki toplam dolu sütun sayisini döndürür."""
        return cls._sheet.max_column